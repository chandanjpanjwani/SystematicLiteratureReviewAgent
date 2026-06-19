#!/usr/bin/env python3
"""
Weekly literature monitor for the SLR on platform strategies in biotech.

Searches OpenAlex and EuropePMC for papers published since the last run,
deduplicates against the known corpus, appends new articles to
output/monitor_new_articles.jsonl, updates output/monitor_state.json,
and emails a digest to the configured recipient.

Usage:
    python3 scripts/weekly_monitor.py
    python3 scripts/weekly_monitor.py --dry-run
    python3 scripts/weekly_monitor.py --from-date 2026-01-01 --to-date 2026-06-19

Environment variables:
    SMTP_PASSWORD   Password for chpa00004@stud.uni-saarland.de (required for email)
"""

import argparse
import json
import os
import smtplib
import ssl
import time
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
STATE_FILE = BASE / "output" / "monitor_state.json"
NEW_ARTICLES_FILE = BASE / "output" / "monitor_new_articles.jsonl"
QUERIES_FILE = BASE / "data" / "01_raw" / "queries.json"
CORPUS_EXCEL = BASE / "output" / "excel" / "Corpus.xlsx"

SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = "cpanjwani00003@gmail.com"
EMAIL_TO = ["chpa00004@stud.uni-saarland.de", "nicolas.kiefer@uni-saarland.de"]
OPENALEX_MAILTO = "chpa00004@stud.uni-saarland.de"


# ---------------------------------------------------------------------------
# State management
# ---------------------------------------------------------------------------

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"last_run": None, "known_dois": [], "known_titles": []}


def save_state(state: dict):
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2))


def load_known_ids(state: dict) -> tuple[set, set]:
    """Return (known_dois, known_title_tokens) from state, or seed from Corpus.xlsx."""
    known_dois = set(d.lower() for d in state.get("known_dois", []) if d)
    known_titles = set(t.lower()[:60] for t in state.get("known_titles", []) if t)

    if not known_dois and CORPUS_EXCEL.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(CORPUS_EXCEL, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            doi_col = next((i for i, h in enumerate(headers) if h and "doi" in str(h).lower()), None)
            title_col = next((i for i, h in enumerate(headers) if h and "title" in str(h).lower()), None)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if doi_col is not None and row[doi_col]:
                    known_dois.add(str(row[doi_col]).lower())
                if title_col is not None and row[title_col]:
                    known_titles.add(str(row[title_col]).lower()[:60])
            wb.close()
            print(f"Seeded {len(known_dois)} DOIs and {len(known_titles)} titles from Corpus.xlsx")
        except Exception as e:
            print(f"Warning: could not read Corpus.xlsx for seeding: {e}")

    return known_dois, known_titles


def is_duplicate(paper: dict, known_dois: set, known_titles: set) -> bool:
    doi = (paper.get("doi") or "").lower()
    if doi and doi in known_dois:
        return True
    title = (paper.get("title") or "").lower()[:60]
    if title and title in known_titles:
        return True
    return False


# ---------------------------------------------------------------------------
# OpenAlex search (date-filtered)
# ---------------------------------------------------------------------------

def search_openalex(from_date: str, to_date: str) -> list[dict]:
    with open(QUERIES_FILE) as f:
        cfg = json.load(f)["openalex"]

    # Strip existing date filter and inject our window
    base_filter = cfg.get("filter", "")
    base_filter = ",".join(
        p for p in base_filter.split(",")
        if not p.strip().startswith("from_publication_date")
    )
    date_filter = f"from_publication_date:{from_date},to_publication_date:{to_date}"
    full_filter = f"{base_filter},{date_filter}"

    results = []
    cursor = "*"
    per_page = 200

    while cursor:
        params = {
            "filter": full_filter,
            "per-page": per_page,
            "cursor": cursor,
            "mailto": OPENALEX_MAILTO,
        }
        url = f"https://api.openalex.org/works?{urllib.parse.urlencode(params)}"
        req = urllib.request.Request(url, headers={"User-Agent": f"slr-monitor ({OPENALEX_MAILTO})"})
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                page = json.loads(resp.read())
        except Exception as e:
            print(f"OpenAlex error: {e}")
            break

        for w in page.get("results", []):
            abstract = None
            inv = w.get("abstract_inverted_index")
            if inv:
                positions = {}
                for word, idxs in inv.items():
                    for i in idxs:
                        positions[i] = word
                abstract = " ".join(positions[i] for i in sorted(positions))
            results.append({
                "id": w.get("id"),
                "doi": (w.get("doi") or "").replace("https://doi.org/", "") or None,
                "title": w.get("title"),
                "abstract": abstract,
                "authors": [a.get("author", {}).get("display_name") for a in (w.get("authorships") or []) if a.get("author", {}).get("display_name")],
                "year": w.get("publication_year"),
                "venue": ((w.get("primary_location") or {}).get("source") or {}).get("display_name"),
                "oa_url": (w.get("open_access") or {}).get("oa_url"),
                "source": "openalex",
            })

        cursor = page.get("meta", {}).get("next_cursor")
        time.sleep(0.1)

    print(f"OpenAlex: {len(results)} results for {from_date} → {to_date}")
    return results


# ---------------------------------------------------------------------------
# EuropePMC search (date-filtered)
# ---------------------------------------------------------------------------

def search_epmc(from_date: str, to_date: str) -> list[dict]:
    with open(QUERIES_FILE) as f:
        cfg = json.load(f)["europepmc"]

    from_year = from_date[:4]
    to_year = to_date[:4]

    # Replace the year range in the existing query
    base_query = cfg["query"]
    base_query = base_query.replace(
        f"(PUB_YEAR:[2005 TO 2026])",
        f"(PUB_YEAR:[{from_year} TO {to_year}])"
    )
    # Also add full date filter via FIRST_PDATE
    date_query = f'({base_query}) AND (FIRST_PDATE:[{from_date} TO {to_date}])'

    results = []
    cursor = "*"
    page_size = 100

    while True:
        params = {
            "query": date_query,
            "format": "json",
            "resultType": "core",
            "pageSize": page_size,
            "cursorMark": cursor,
        }
        url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/search?{urllib.parse.urlencode(params)}"
        req = urllib.request.Request(url, headers={"User-Agent": "slr-monitor"})
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                page = json.loads(resp.read())
        except Exception as e:
            print(f"EuropePMC error: {e}")
            break

        for r in (page.get("resultList") or {}).get("result") or []:
            results.append({
                "id": f"EPMC:{r.get('id')}:{r.get('source')}",
                "doi": r.get("doi"),
                "pmid": r.get("pmid"),
                "title": r.get("title"),
                "abstract": r.get("abstractText"),
                "authors": [a.get("fullName") for a in (r.get("authorList") or {}).get("author", []) if a.get("fullName")],
                "year": int(r.get("pubYear")) if r.get("pubYear") else None,
                "venue": r.get("journalTitle"),
                "is_open_access": r.get("isOpenAccess") == "Y",
                "source": "europepmc",
            })

        next_cursor = page.get("nextCursorMark")
        if not next_cursor or next_cursor == cursor:
            break
        cursor = next_cursor
        time.sleep(0.2)

    print(f"EuropePMC: {len(results)} results for {from_date} → {to_date}")
    return results


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def send_email(new_papers: list[dict], from_date: str, to_date: str, dry_run: bool):
    smtp_password = os.environ.get("SMTP_PASSWORD", "")
    if not smtp_password and not dry_run:
        print("Warning: SMTP_PASSWORD not set — skipping email")
        return

    subject = f"SLR Weekly Monitor: {len(new_papers)} new paper(s) [{from_date} → {to_date}]"

    lines = [
        f"Weekly literature monitor — platform strategies in biotech",
        f"Search window: {from_date} to {to_date}",
        f"New papers found: {len(new_papers)}",
        "",
    ]

    if not new_papers:
        lines.append("No new papers found this week.")
    else:
        for i, p in enumerate(new_papers, 1):
            title = p.get("title") or "(no title)"
            authors = ", ".join((p.get("authors") or [])[:3])
            if len(p.get("authors") or []) > 3:
                authors += " et al."
            year = p.get("year") or ""
            venue = p.get("venue") or ""
            doi = p.get("doi") or ""
            url = f"https://doi.org/{doi}" if doi else (p.get("oa_url") or p.get("id") or "")
            lines += [
                f"{i}. {title}",
                f"   {authors} ({year}) — {venue}",
                f"   {url}",
                "",
            ]

    body = "\n".join(lines)

    recipients = EMAIL_TO if isinstance(EMAIL_TO, list) else [EMAIL_TO]

    if dry_run:
        print("\n--- EMAIL PREVIEW ---")
        print(f"To: {', '.join(recipients)}")
        print(f"Subject: {subject}")
        print(body)
        print("--- END PREVIEW ---\n")
        return

    recipients = EMAIL_TO if isinstance(EMAIL_TO, list) else [EMAIL_TO]

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(recipients)
    msg.attach(MIMEText(body, "plain"))

    ctx = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls(context=ctx)
        server.login(SMTP_USER, smtp_password)
        server.sendmail(SMTP_USER, recipients, msg.as_string())
    print(f"Email sent to {', '.join(recipients)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run", action="store_true", help="Print results without writing files or sending email")
    p.add_argument("--from-date", help="Override start date (YYYY-MM-DD)")
    p.add_argument("--to-date", help="Override end date (YYYY-MM-DD)")
    args = p.parse_args()

    today = date.today().isoformat()
    state = load_state()

    from_date = args.from_date or state.get("last_run") or (date.today() - timedelta(days=7)).isoformat()
    to_date = args.to_date or today

    print(f"Monitor run: {from_date} → {to_date}")

    known_dois, known_titles = load_known_ids(state)

    # Search both databases
    candidates = search_openalex(from_date, to_date) + search_epmc(from_date, to_date)
    print(f"Total candidates: {len(candidates)}")

    # Deduplicate against corpus and against each other
    new_papers = []
    seen_in_batch: set[str] = set()
    for paper in candidates:
        if is_duplicate(paper, known_dois, known_titles):
            continue
        # Deduplicate within batch by DOI then title
        doi = (paper.get("doi") or "").lower()
        title_key = (paper.get("title") or "").lower()[:60]
        batch_key = doi or title_key
        if batch_key and batch_key in seen_in_batch:
            continue
        if batch_key:
            seen_in_batch.add(batch_key)
        new_papers.append(paper)

    print(f"New papers after deduplication: {len(new_papers)}")

    if not args.dry_run:
        # Append new articles
        NEW_ARTICLES_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(NEW_ARTICLES_FILE, "a") as f:
            for paper in new_papers:
                f.write(json.dumps(paper) + "\n")

        # Update state
        new_dois = [p.get("doi") for p in new_papers if p.get("doi")]
        new_titles = [p.get("title") for p in new_papers if p.get("title")]
        state["last_run"] = today
        state["known_dois"] = list(set(state.get("known_dois", [])) | set(new_dois))
        state["known_titles"] = list(set(state.get("known_titles", [])) | set(new_titles))
        save_state(state)
        print(f"State updated. Total known DOIs: {len(state['known_dois'])}")

    send_email(new_papers, from_date, to_date, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
